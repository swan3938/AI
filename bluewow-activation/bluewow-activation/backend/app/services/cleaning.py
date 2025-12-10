import csv
from io import TextIOWrapper, BytesIO
from openpyxl import load_workbook

def preview_csv(file) -> list:
    rows = []
    wrapper = TextIOWrapper(file, encoding="utf-8")
    reader = csv.DictReader(wrapper)
    for i, row in enumerate(reader):
        rows.append(row)
        if i >= 9:
            break
    return rows

def read_csv_all(path: str) -> list:
    out = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            out.append(row)
    return out

def preview_xlsx(file) -> list:
    data = file.read()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        item = {}
        for j, h in enumerate(headers):
            v = row[j] if j < len(row) else None
            if v is None:
                v = ""
            else:
                v = str(v)
            item[h] = v
        rows.append(item)
        if len(rows) >= 10:
            break
    return rows

def read_xlsx_all(path: str) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        item = {}
        for j, h in enumerate(headers):
            v = row[j] if j < len(row) else None
            if v is None:
                v = ""
            else:
                v = str(v)
            item[h] = v
        out.append(item)
    return out

def read_all(path: str) -> list:
    if path.lower().endswith(".xlsx") or path.lower().endswith(".xls"):
        return read_xlsx_all(path)
    return read_csv_all(path)

def read_csv_stream(file) -> list:
    out = []
    wrapper = TextIOWrapper(file, encoding="utf-8")
    reader = csv.DictReader(wrapper)
    for row in reader:
        out.append(row)
    return out

def read_all_stream(file, filename: str) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        data = file.read()
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        out = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip() if h is not None else "" for h in row]
                continue
            item = {}
            for j, h in enumerate(headers):
                v = row[j] if j < len(row) else None
                if v is None:
                    v = ""
                else:
                    v = str(v)
                item[h] = v
            out.append(item)
        return out
    return read_csv_stream(file)
